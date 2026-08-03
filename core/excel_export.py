"""Excel workbook generation for the combined AHP / TOPSIS / Sensitivity results."""

import io

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.ahp import slider_to_aij
from core.data import CRITERIA, EXPERTS, N, TECHNIQUES
from core.topsis import run_sensitivity, run_topsis


def style_header(ws, row):
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def add_dataframe(ws, df, start_row=1, start_col=1):
    for c, column in enumerate(df.columns, start_col):
        ws.cell(start_row, c, column)
    style_header(ws, start_row)

    for r, record in enumerate(df.itertuples(index=False), start_row + 1):
        for c, value in enumerate(record, start_col):
            ws.cell(r, c, value)


def create_results_workbook(saved_experts):
    wb = Workbook()
    wb.remove(wb.active)

    expert_weights = []

    for expert in EXPERTS:
        if expert not in saved_experts:
            continue

        data = saved_experts[expert]
        ws = wb.create_sheet(expert.replace(" ", "_"))
        ws["A1"] = f"AHP responses — {expert}"
        ws["A1"].font = Font(bold=True, size=14)

        headers = ["Left Criterion", "Right Criterion", "AHP Value"]
        for col, header in enumerate(headers, 1):
            ws.cell(3, col, header)
        style_header(ws, 3)

        row = 4
        for (i, j), pos in data["positions"].items():
            ws.cell(row, 1, CRITERIA[i]["name"])
            ws.cell(row, 2, CRITERIA[j]["name"])
            ws.cell(row, 3, slider_to_aij(pos))
            row += 1

        matrix = data["matrix"]
        weights = data["weights"]
        cr = data["cr"]
        expert_weights.append(weights)

        row += 1
        ws.cell(row, 1, "Pairwise Comparison Matrix")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        for c, criterion in enumerate(CRITERIA, 2):
            ws.cell(row, c, criterion["id"])
        for r in range(N):
            ws.cell(row + 1 + r, 1, CRITERIA[r]["id"])
            for c in range(N):
                ws.cell(row + 1 + r, 2 + c, float(matrix[r, c]))

        row += N + 3
        ws.cell(row, 1, "Criterion")
        ws.cell(row, 2, "Weight")
        style_header(ws, row)
        for i, criterion in enumerate(CRITERIA, row + 1):
            ws.cell(i, 1, criterion["name"])
            ws.cell(i, 2, float(weights[i - row - 1]))

        ws.cell(row + N + 2, 1, "Consistency Ratio")
        ws.cell(row + N + 2, 2, float(cr))

    if not expert_weights:
        return None

    # Aggregate experts using arithmetic mean, then normalise
    final_weights = np.mean(np.vstack(expert_weights), axis=0)
    final_weights = final_weights / final_weights.sum()

    ws = wb.create_sheet("AHP_Weights")
    ahp_df = pd.DataFrame({
        "Criterion ID": [c["id"] for c in CRITERIA],
        "Criterion": [c["name"] for c in CRITERIA],
        "Final Weight": final_weights,
    })
    add_dataframe(ws, ahp_df)

    topsis_df = run_topsis(final_weights)
    ws = wb.create_sheet("TOPSIS_Results")
    add_dataframe(ws, topsis_df)

    sensitivity_df = run_sensitivity(final_weights, topsis_df)
    ws = wb.create_sheet("Sensitivity_Analysis")
    add_dataframe(ws, sensitivity_df)

    ws = wb.create_sheet("Decision_Matrix")
    decision_df = pd.DataFrame(
        [[x[0], x[1], *x[2]] for x in TECHNIQUES],
        columns=["Alternative", "Technique", *[c["id"] for c in CRITERIA]],
    )
    add_dataframe(ws, decision_df)

    # Make columns readable
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            ws.column_dimensions[column_cells[0].column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), final_weights, topsis_df, sensitivity_df
